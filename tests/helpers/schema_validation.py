"""Test helpers for validating artifact schemas."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Iterable, Sequence

from jsonschema import Draft7Validator
from openpyxl import load_workbook


def assert_json_schema(payload_path: Path, schema_path: Path) -> None:
    """Validate ``payload_path`` against the JSON schema at ``schema_path``."""

    data = json.loads(payload_path.read_text(encoding="utf-8"))
    schema = json.loads(schema_path.read_text(encoding="utf-8"))
    validator = Draft7Validator(schema)
    errors = sorted(validator.iter_errors(data), key=lambda error: error.path)
    if errors:
        details = "\n".join(
            f"{'/'.join(map(str, error.path)) or '<root>'}: {error.message}" for error in errors
        )
        raise AssertionError(
            f"Payload {payload_path} does not satisfy schema {schema_path}:\n{details}"
        )


def _normalise_header(values: Iterable[object]) -> list[str]:
    header: list[str] = []
    for value in values:
        header.append("" if value is None else str(value))
    return header


def assert_excel_columns(path: Path, sheet: str, expected: Sequence[str]) -> None:
    """Validate that ``sheet`` in ``path`` exposes the ``expected`` column headers."""

    workbook = load_workbook(path)
    if sheet not in workbook.sheetnames:
        raise AssertionError(f"Workbook {path} does not contain sheet {sheet!r}")
    worksheet = workbook[sheet]
    rows = worksheet.iter_rows(max_row=1)
    try:
        header_row = next(rows)
    except StopIteration as exc:  # pragma: no cover - defensive guard
        raise AssertionError(f"Sheet {sheet!r} in {path} is empty") from exc
    header = _normalise_header(cell.value for cell in header_row)
    expected_list = [str(column) for column in expected]
    if header != expected_list:
        raise AssertionError(
            "Unexpected columns for sheet "
            f"{sheet!r} in {path}: expected {expected_list} but received {header}"
        )
